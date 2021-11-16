import json
from json.decoder import JSONDecodeError
from functions_pkg.db_functions import MySQLConnection

from PyQt5 import QtWidgets
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDateTime, QRegExp, QDate
from PyQt5.QtWidgets import QWidget, QApplication, QComboBox, QLabel, QFileDialog, QInputDialog, QMessageBox, \
    QPushButton, QProgressDialog
from openpyxl import load_workbook

import functions_pkg.functions as func
from GLOBAL_VARS import *
from equipment_pkg.ui_equipment_add_file import Ui_Form
from functions_pkg.send_get_request import GetRequest

COMBOBOX_VALUES = ["",
                   "Номер регистрационной карточки",
                   "Область измерений по МИ 2314-2006",
                   "Номер (номера) отдела",
                   "ФИО ответственного",
                   "Помещение",
                   "Регистрационный номер типа СИ",
                   "Межповерочный интервал",
                   "Наименование типа",
                   "Тип",
                   "Модификация",
                   "Заводской номер",
                   "Инвентарный номер",
                   "Изготовитель",
                   "Год выпуска",
                   "Год введения в эксплуатацию",
                   "Диапазон измерений",
                   "Погрешность",
                   "Класс точности",
                   "Прочие характеристики",
                   "Наличие паспорта (да/нет, 1/0, истина/ложь, true/false)",
                   "Наличие РЭ (да/нет, 1/0, истина/ложь, true/false)",
                   "Наличие МП (да/нет, 1/0, истина/ложь, true/false)",
                   "Встроенное программное обеспечение",
                   "Внешнее программное обеспечение",
                   "Назначение",
                   "Допущенный персонал",
                   "Собственник",
                   "Номер и дата договора (если СИ не в собственности)",
                   "Периодичность технического обслуживания",
                   "Содержание технического обслуживания",
                   "1. Дата проведения ТО (в формате 31.01.2021)",
                   "1. Сотрудник, проводивший ТО",
                   "2. Дата проведения ТО (в формате 31.01.2021)",
                   "2. Сотрудник, проводивший ТО",
                   "3. Дата проведения ТО (в формате 31.01.2021)",
                   "3. Сотрудник, проводивший ТО",
                   "4. Дата проведения ТО (в формате 31.01.2021)",
                   "4. Сотрудник, проводивший ТО",
                   "5. Дата проведения ТО (в формате 31.01.2021)",
                   "5. Сотрудник, проводивший ТО",
                   "1. Номер свидетельства",
                   "1. Дата поверки (в формате 31.01.2021)",
                   "1. Действительна до (в формате 31.01.2021)",
                   "1. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "1. Организация-поверитель",
                   "1. СИ как эталон - номер в перечне",
                   "2. Номер свидетельства",
                   "2. Дата поверки (в формате 31.01.2021)",
                   "2. Действительна до (в формате 31.01.2021)",
                   "2. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "2. Организация-поверитель",
                   "2. СИ как эталон - номер в перечне",
                   "3. Номер свидетельства",
                   "3. Дата поверки (в формате 31.01.2021)",
                   "3. Действительна до (в формате 31.01.2021)",
                   "3. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "3. Организация-поверитель",
                   "3. СИ как эталон - номер в перечне",
                   "4. Номер свидетельства",
                   "4. Дата поверки (в формате 31.01.2021)",
                   "4. Действительна до (в формате 31.01.2021)",
                   "4. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "4. Организация-поверитель",
                   "4. СИ как эталон - номер в перечне",
                   "5. Номер свидетельства",
                   "5. Дата поверки (в формате 31.01.2021)",
                   "5. Действительна до (в формате 31.01.2021)",
                   "5. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "5. Организация-поверитель",
                   "5. СИ как эталон - номер в перечне",
                   "6. Номер свидетельства",
                   "6. Дата поверки (в формате 31.01.2021)",
                   "6. Действительна до (в формате 31.01.2021)",
                   "6. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "6. Организация-поверитель",
                   "6. СИ как эталон - номер в перечне",
                   "7. Номер свидетельства",
                   "7. Дата поверки (в формате 31.01.2021)",
                   "7. Действительна до (в формате 31.01.2021)",
                   "7. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "7. Организация-поверитель",
                   "7. СИ как эталон - номер в перечне",
                   "8. Номер свидетельства",
                   "8. Дата поверки (в формате 31.01.2021)",
                   "8. Действительна до (в формате 31.01.2021)",
                   "8. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "8. Организация-поверитель",
                   "8. СИ как эталон - номер в перечне",
                   "9. Номер свидетельства",
                   "9. Дата поверки (в формате 31.01.2021)",
                   "9. Действительна до (в формате 31.01.2021)",
                   "9. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "9. Организация-поверитель",
                   "9. СИ как эталон - номер в перечне",
                   "10. Номер свидетельства",
                   "10. Дата поверки (в формате 31.01.2021)",
                   "10. Действительна до (в формате 31.01.2021)",
                   "10. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)",
                   "10. Организация-поверитель",
                   "10. СИ как эталон - номер в перечне"
                   ]

COLUMN_NAMES = ["A(1)", "B(2)", "C(3)", "D(4)", "E(5)", "F(6)", "G(7)", "H(8)", "I(9)", "J(10)", "K(11)", "L(12)",
                "M(13)", "N(14)", "O(15)", "P(16)", "Q(17)", "R(18)", "S(19)", "T(20)", "U(21)", "V(22)", "W(23)",
                "X(24)", "Y(25)", "Z(26)", "AA(27)", "AB(28)", "AC(29)", "AD(30)", "AE(31)", "AF(32)", "AG(33)",
                "AH(34)", "AI(35)", "AJ(36)", "AK(37)", "AL(38)", "AM(39)", "AN(40)", "AO(41)", "AP(42)", "AQ(43)",
                "AR(44)", "AS(45)", "AT(46)", "AU(47)", "AV(48)", "AW(49)", "AX(50)", "AY(51)", "AZ(52)", "BA(53)",
                "BB(54)", "BC(55)", "BD(56)", "BE(57)", "BF(58)", "BG(59)", "BH(60)", "BI(61)", "BJ(62)", "BK(63)",
                "BL(64)", "BM(65)", "BN(66)", "BO(67)", "BP(68)", "BQ(69)", "BR(70)", "BS(71)", "BT(72)", "BU(73)",
                "BV(74)", "BW(75)", "BX(76)", "BY(77)", "BZ(78)", "CA(79)", "CB(80)", "CC(81)", "CD(82)", "CE(83)",
                "CF(84)", "CG(85)", "CH(86)", "CI(87)", "CJ(88)", "CK(89)", "CL(90)", "CM(91)", "CN(92)", "CO(93)",
                "CP(94)", "CQ(95)", "CR(96)", "CS(97)", "CT(98)", "CU(99)", "CV(100)", "CW(101)", "CX(102)", "CY(103)",
                "CZ(104)"]


class SearchThread(QThread):
    msg_signal = pyqtSignal(str)

    def __init__(self, parent=None):
        QThread.__init__(self, parent)
        self.url = ""
        self.is_running = True

    def run(self):
        if self.is_running:
            self.sleep(1)
            print("thread running")
            print(f" {self.url}")
            resp = GetRequest.getRequest(self.url)
            print(f"  {resp}")
            print("    thread stopped")
            self.msg_signal.emit(resp)
        else:
            self.msg_signal.emit("stop")


class EquipmentImportFileWidget(QWidget):
    def __init__(self):
        super(EquipmentImportFileWidget, self).__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.search_thread = SearchThread()

        self.column_count = 0

        self.worker_dict = func.get_workers()['worker_dict']
        self.room_dict = func.get_rooms()['room_dict']

        self.combobox_list = list()
        self.combobox_items_dict = dict()
        self.final_dict = dict()

        self.error_row_numbers = list()

        self.mit = self.mit_search = dict()

        mis = func.get_mis()
        self.mi_dict = mis['mi_dict']
        self.set_of_card_numbers = set()
        self.dep_dict = func.get_departments()['dep_dict']

        for column_name in COLUMN_NAMES:
            self.combobox_items_dict[column_name] = ""

        # self.ui.pushButton_add_column.clicked.connect(lambda: self._add_column())
        self.ui.pushButton_add_columns.clicked.connect(lambda: self._add_columns(self.ui.spinBox_add_columns.value()))
        self.ui.pushButton_file_select.clicked.connect(self._file_select)
        self.ui.pushButton_import_start.clicked.connect(self._verify_columns)
        self.search_thread.msg_signal.connect(self._on_getting_resp, Qt.QueuedConnection)

        self._add_column()

    # ДОБАВЛЕНИЕ КОЛОНОК ПРИ КЛИКЕ НА КНОПКУ "ДОБАВИТЬ"
    def _add_columns(self, col_number=1):
        for i in range(col_number):
            self._add_column()

    # ДОБАВЛЕНИЕ ОДНОЙ КОЛОНКИ
    def _add_column(self):
        if self.column_count < len(COLUMN_NAMES):
            cur_label_name = COLUMN_NAMES[self.column_count]
            # СОЗДАЕМ НАДПИСЬ
            label = QLabel(self.ui.scrollAreaWidgetContents)
            label.setText(f"Колонка {cur_label_name}")
            # СОЗДАЕМ ПОЛЕ ВЫБОРА С ИМЕНЕМ КОЛОНКИ
            comboBox = QComboBox(self.ui.scrollAreaWidgetContents)
            comboBox.setObjectName(cur_label_name)

            # ДОБАЛЯЕМ НАДПИСЬ И КОМБОБОКС В ФОРМУ
            self.ui.formLayout.setWidget(0 + self.column_count, QtWidgets.QFormLayout.LabelRole, label)
            self.ui.formLayout.setWidget(0 + self.column_count, QtWidgets.QFormLayout.FieldRole, comboBox)

            # ЗАДАЕМ ПОВЕДЕНИЕ ПРИ ВЫБОРЕ ЗНАЧЕНИЯ В КОМБОБОКСЕ (ПЕРЕСЧЕТ ЗНАЧЕНИЙ ВСЕХ ДРУГИХ КОМБОБОКСОВ)
            comboBox.textActivated.connect(lambda: self._add_combobox_items(comboBox))
            self.column_count = self.column_count + 1
            # ДОБАВЛЯЕМ КОМБОБОКС В СПИСОК КОМБОБОКСОВ
            self.combobox_list.append(comboBox)
            # ЗАПУСКАЕМ ПЕРЕСЧЕТ ЗНАЧЕНИЙ ВСЕХ КОМБОБОКСОВ
            self._add_combobox_items(comboBox)

    # ЗАДАНИЕ ЗНАЧЕНИЙ ДЛЯ ВСЕХ КОМБОБОКСОВ ПРИ ДОБАВЛЕНИИ ИЛИ ИЗМЕНЕНИИ ЗНАЧЕНИЯ ДРУГОГО ТЕКУЩЕГО КОМБОБОКСА
    def _add_combobox_items(self, cur_combobox):
        self.combobox_items_dict[cur_combobox.objectName()] = cur_combobox.currentText()

        for combobox in self.combobox_list:
            combobox.clear()
            for column_name in COMBOBOX_VALUES:
                combobox.addItem(column_name)
            separators = [31, 34, 37, 40, 43, 46, 53, 60, 67, 74, 81, 88, 95, 102, 109, 116]
            for index in separators:
                combobox.insertSeparator(index)
            for column_name in self.combobox_items_dict:
                if self.combobox_items_dict[column_name] != "" and column_name != combobox.objectName():
                    combobox.removeItem(combobox.findText(self.combobox_items_dict[column_name]))

            combobox.setCurrentText(self.combobox_items_dict[combobox.objectName()])

    # ВЫБОР ИМПОРТИРУЕМОГО ФАЙЛА
    def _file_select(self):
        dir_name = str(SETTINGS.value("paths/cometa_path"))
        file_path = QFileDialog.getOpenFileName(caption="Выберите файл для импорта", directory=dir_name,
                                                filter="Excel (*.xlsx)")[0]
        self.ui.lineEdit_file_path.setText(file_path)
        self.wb = load_workbook(filename=file_path)
        self.ui.spinBox_end_row.setValue(self.wb.active.max_row)
        self.ui.spinBox_end_row.setMaximum(self.wb.active.max_row)
        self.ui.spinBox_start_row.setMaximum(self.wb.active.max_row)
        self.ui.spinBox_add_columns.setValue(self.wb.active.max_column - 1)

    # ПРОВЕРКА ДОСТАТОЧНОСТИ ВЫБРАННЫХ КОЛОНОК ДЛЯ ИМПОРТА
    def _verify_columns(self):
        # rx_svid = QRegExp("^(С|И)\-\S{1,3}\/[0-3][0-9]\-[0-1][0-9]\-20[2-5][0-9]\/\d{8,10}$")
        # rx_svid.setCaseSensitivity(False)
        # if rx_svid.indexIn(dialog.textValue()) == 0:
        #     dialog.setLabelText(f"Номер свидетельства о поверке")
        #     self.eq_type = "vri_id"

        number = inv_number = reestr = title = cert_number = mieta_number = False

        column_names = set()
        for column_name in self.combobox_items_dict:
            if self.combobox_items_dict[column_name]:
                column_names.add(self.combobox_items_dict[column_name])
            if self.combobox_items_dict[column_name] == "Заводской номер":
                number = True
            elif self.combobox_items_dict[column_name] == "Инвентарный номер":
                inv_number = True
            elif self.combobox_items_dict[column_name] == "Регистрационный номер типа СИ":
                reestr = True
            elif self.combobox_items_dict[column_name] == "Наименование типа":
                title = True
            elif "Номер свидетельства" in self.combobox_items_dict[column_name]:
                cert_number = True
            elif "СИ как эталон" in self.combobox_items_dict[column_name]:
                mieta_number = True
        # print(column_names)

        for i in range(1, 11):
            if f"{i}. Номер свидетельства" not in column_names:
                if f"{i}. Дата поверки (в формате 31.01.2021)" in column_names:
                    QMessageBox.critical(self, "Ошибка",
                                         f"Для колонки \"{i}. Дата поверки (в формате 31.01.2021)\" отсутствует "
                                         f"колонка \"{i}. Номер свидетельства\".\n"
                                         f"Необходимо поменять номер группы!")
                    return
                elif f"{i}. Действительна до (в формате 31.01.2021)" in column_names:
                    QMessageBox.critical(self, "Ошибка",
                                         f"Для колонки \"{i}. Действительна до (в формате 31.01.2021)\" "
                                         f"отсутствует колонка \"{i}. Номер свидетельства\".\n"
                                         f"Необходимо поменять номер группы!")
                    return
                elif f"{i}. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)" in column_names:
                    QMessageBox.critical(self, "Ошибка",
                                         f"Для колонки \"{i}. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, "
                                         f"true/false)\" отсутствует колонка \"{i}. Номер свидетельства\".\n"
                                         f"Необходимо поменять номер группы!")
                    return
                elif f"{i}. Организация-поверитель" in column_names:
                    QMessageBox.critical(self, "Ошибка",
                                         f"Для колонки \"{i}. Организация-поверитель\" отсутствует колонка "
                                         f"\"{i}. Номер свидетельства\".\n"
                                         f"Необходимо поменять номер группы!")
                    return

        if not cert_number and not mieta_number:
            if reestr or title:
                if not number and not inv_number:
                    QMessageBox.critical(self, "Ошибка",
                                         f"Необходимо также назначить заводской или инвентарный номер")
                    return

            else:
                QMessageBox.critical(self, "Ошибка",
                                     f"Для успешного импорта необходимо назначить хотя бы одну из следующих "
                                     f"комбинаций колонок:\n"
                                     f"- номер свидетельства;\n"
                                     f"- СИ как эталон - номер в перечне\n"
                                     f"- регистрационный номер типа СИ + заводской (инвентарный) номер\n"
                                     f"- наименование типа + заводской (инвентарный) номер.")
                return

        self._import_start()

    def _import_start(self):
        if not self.ui.lineEdit_file_path.text():
            QMessageBox.critical(self, "Ошибка",
                                 f"Необходимо выбрать файл для импорта!")
            return

        self.error_row_numbers.clear()
        self.set_of_card_numbers.clear()
        sheet = self.wb.active
        start_row = self.ui.spinBox_start_row.value()
        end_row = self.ui.spinBox_end_row.value()
        count = end_row - start_row + 1

        self.dialog = QProgressDialog(self)
        self.dialog.setAutoClose(False)
        self.dialog.setAutoReset(False)
        self.dialog.setWindowTitle("ОЖИДАЙТЕ! Идет импорт!")
        self.dialog.setRange(0, 100)
        self.dialog.setWindowModality(Qt.WindowModal)
        self.dialog.setCancelButton(None)
        self.dialog.resize(350, 100)
        self.dialog.show()
        self.dialog.setLabelText(f"Импортировано 0 из {count}")
        self.dialog.setValue(0)

        for i in range(start_row, end_row + 1):
            self.final_dict[i] = dict()

            reg_card_number = measure_code = department = responsiblePerson = room = reestr = MPI \
                = title = type = modification = number = inv_number = manufacturer = manuf_year \
                = expl_year = diapazon = PG = KT = other_characteristics = has_pasport = has_manual \
                = has_verif_method = software_inner = software_outer = purpose = personal = owner \
                = owner_contract = period_TO = content_TO = TO_date_1 = TO_date_2 = TO_date_3 = TO_date_4 \
                = TO_date_5 = TO_worker_1 = TO_worker_2 = TO_worker_3 = TO_worker_4 = TO_worker_5 \
                = mieta_number_1 = mieta_number_2 = mieta_number_3 = mieta_number_4 = mieta_number_5 \
                = mieta_number_6 = mieta_number_7 = mieta_number_8 = mieta_number_9 = mieta_number_10 = ""

            mieta_numbers = list()
            vrf_dates = list()
            valid_dates = list()
            cert_numbers = list()
            results = list()
            organizations = list()

            for k in range(10):
                mieta_numbers.append("")
                vrf_dates.append("")
                valid_dates.append("")
                cert_numbers.append("")
                results.append("")
                organizations.append("")

            for column_name in self.combobox_items_dict:
                column_letter = column_name.split("(")[0]
                if str(sheet[f"{column_letter}{i}"].value) != "None":

                    if self.combobox_items_dict[column_name] == "Номер регистрационной карточки":
                        reg_card_number = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Область измерений по МИ 2314-2006":
                        measure_code = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Номер (номера) отдела":
                        department = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "ФИО ответственного":
                        responsiblePerson = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Помещение":
                        room = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Регистрационный номер типа СИ":
                        reestr = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Межповерочный интервал":
                        MPI = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Наименование типа":
                        title = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Тип":
                        type = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Модификация":
                        modification = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Заводской номер":
                        number = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Инвентарный номер":
                        inv_number = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Изготовитель":
                        manufacturer = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Год выпуска":
                        manuf_year = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Год введения в эксплуатацию":
                        expl_year = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Диапазон измерений":
                        diapazon = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Погрешность":
                        PG = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Класс точности":
                        KT = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Прочие характеристики":
                        other_characteristics = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[
                        column_name] == "Наличие паспорта (да/нет, 1/0, истина/ложь, true/false)":
                        has_pasport = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Наличие РЭ (да/нет, 1/0, истина/ложь, true/false)":
                        has_manual = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Наличие МП (да/нет, 1/0, истина/ложь, true/false)":
                        has_verif_method = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Встроенное программное обеспечение":
                        software_inner = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Внешнее программное обеспечение":
                        software_outer = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Назначение":
                        purpose = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Допущенный персонал":
                        personal = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Собственник":
                        owner = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Номер и дата договора (если СИ не в собственности)":
                        owner_contract = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Периодичность технического обслуживания":
                        period_TO = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "Содержание технического обслуживания":
                        content_TO = str(sheet[f"{column_letter}{i}"].value)

                    # for k in range(1, 11):

                    elif self.combobox_items_dict[column_name] == "1. Дата проведения ТО (в формате 31.01.2021)":
                        TO_date_1 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "2. Дата проведения ТО (в формате 31.01.2021)":
                        TO_date_2 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "3. Дата проведения ТО (в формате 31.01.2021)":
                        TO_date_3 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "4. Дата проведения ТО (в формате 31.01.2021)":
                        TO_date_4 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "5. Дата проведения ТО (в формате 31.01.2021)":
                        TO_date_5 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "1. Сотрудник, проводивший ТО":
                        TO_worker_1 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "2. Сотрудник, проводивший ТО":
                        TO_worker_2 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "3. Сотрудник, проводивший ТО":
                        TO_worker_3 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "4. Сотрудник, проводивший ТО":
                        TO_worker_4 = str(sheet[f"{column_letter}{i}"].value)
                    elif self.combobox_items_dict[column_name] == "5. Сотрудник, проводивший ТО":
                        TO_worker_5 = str(sheet[f"{column_letter}{i}"].value)

                    for k in range(1, 11):
                        if self.combobox_items_dict[column_name] == f"{k}. СИ как эталон - номер в перечне":
                            mieta_numbers[k - 1] = str(sheet[f"{column_letter}{i}"].value)

                        elif self.combobox_items_dict[column_name] == f"{k}. Дата поверки (в формате 31.01.2021)":
                            if sheet[f"{column_letter}{i}"].is_date:
                                temp_date = sheet[f"{column_letter}{i}"].value
                                vrf_dates[k - 1] = f"{temp_date:%d.%m.%Y}"
                            else:
                                vrf_dates[k - 1] = str(sheet[f"{column_letter}{i}"].value)

                        elif self.combobox_items_dict[column_name] == f"{k}. Действительна до (в формате 31.01.2021)":
                            if sheet[f"{column_letter}{i}"].is_date:
                                temp_date = sheet[f"{column_letter}{i}"].value
                                valid_dates[k - 1] = f"{temp_date:%d.%m.%Y}"
                            else:
                                valid_dates[k - 1] = str(sheet[f"{column_letter}{i}"].value)

                        elif self.combobox_items_dict[column_name] == f"{k}. Номер свидетельства":
                            cert_numbers[k - 1] = str(sheet[f"{column_letter}{i}"].value)

                        elif self.combobox_items_dict[
                            column_name] == f"{k}. Пригодность (годен/брак, да/нет, 1/0, истина/ложь, true/false)":
                            results[k - 1] = str(sheet[f"{column_letter}{i}"].value)

                        elif self.combobox_items_dict[column_name] == f"{k}. Организация-поверитель":
                            organizations[k - 1] = str(sheet[f"{column_letter}{i}"].value)

            self.final_dict[i] = dict()
            self.final_dict[i]['reg_card_number'] = self._format_string(reg_card_number)
            self.final_dict[i]['measure_code'] = self._format_string(measure_code)
            self.final_dict[i]['department'] = self._format_string(department)
            self.final_dict[i]['responsiblePerson'] = self._format_string(responsiblePerson)
            self.final_dict[i]['room'] = self._format_string(room)
            self.final_dict[i]['reestr'] = self._format_string(reestr)
            self.final_dict[i]['MPI'] = self._format_string(MPI)
            self.final_dict[i]['title'] = self._format_string(title)
            self.final_dict[i]['type'] = self._format_string(type)
            self.final_dict[i]['modification'] = self._format_string(modification)
            self.final_dict[i]['number'] = self._format_string(number)
            self.final_dict[i]['inv_number'] = self._format_string(inv_number)
            self.final_dict[i]['manufacturer'] = self._format_string(manufacturer)
            self.final_dict[i]['manuf_year'] = self._format_string(manuf_year)
            self.final_dict[i]['expl_year'] = self._format_string(expl_year)
            self.final_dict[i]['diapazon'] = self._format_string(diapazon)
            self.final_dict[i]['PG'] = self._format_string(PG)
            self.final_dict[i]['KT'] = self._format_string(KT)
            self.final_dict[i]['other_characteristics'] = self._format_string(other_characteristics)
            self.final_dict[i]['has_pasport'] = self._format_string(has_pasport)
            self.final_dict[i]['has_manual'] = self._format_string(has_manual)
            self.final_dict[i]['has_verif_method'] = self._format_string(has_verif_method)
            self.final_dict[i]['software_inner'] = self._format_string(software_inner)
            self.final_dict[i]['software_outer'] = self._format_string(software_outer)
            self.final_dict[i]['purpose'] = self._format_string(purpose)
            self.final_dict[i]['personal'] = self._format_string(personal)
            self.final_dict[i]['owner'] = self._format_string(owner)
            self.final_dict[i]['owner_contract'] = self._format_string(owner_contract)
            self.final_dict[i]['period_TO'] = self._format_string(period_TO)
            self.final_dict[i]['content_TO'] = self._format_string(content_TO)
            self.final_dict[i]['TO_date_1'] = self._format_string(TO_date_1)
            self.final_dict[i]['TO_date_2'] = self._format_string(TO_date_2)
            self.final_dict[i]['TO_date_3'] = self._format_string(TO_date_3)
            self.final_dict[i]['TO_date_4'] = self._format_string(TO_date_4)
            self.final_dict[i]['TO_date_5'] = self._format_string(TO_date_5)
            self.final_dict[i]['TO_worker_1'] = self._format_string(TO_worker_1)
            self.final_dict[i]['TO_worker_2'] = self._format_string(TO_worker_2)
            self.final_dict[i]['TO_worker_3'] = self._format_string(TO_worker_3)
            self.final_dict[i]['TO_worker_4'] = self._format_string(TO_worker_4)
            self.final_dict[i]['TO_worker_5'] = self._format_string(TO_worker_5)
            self.final_dict[i]['mieta_numbers'] = mieta_numbers
            self.final_dict[i]['vrf_dates'] = vrf_dates
            self.final_dict[i]['valid_dates'] = valid_dates
            self.final_dict[i]['cert_numbers'] = cert_numbers
            self.final_dict[i]['results'] = results
            self.final_dict[i]['organizations'] = organizations
            # print(valid_dates)
            # self.cur_row_number = i
            self._verification_start(i)

            self.dialog.setLabelText(f"Импортировано {i} из {count}")
            self.dialog.setValue(i * 100 / count)
        self.dialog.close()

        QMessageBox.information(self, "Внимание!",
                                f"Строки с номерами: {', '.join(self.error_row_numbers)} не импортировались. "
                                f"Проверьте достаточность данных")

        # self._get_dicts_from_fif()

        # for item in self.final_dict:
        #     print(self.final_dict[item])

    # БЕРЕМ ВСЕ ВОЗМОЖНЫЕ ДАННЫЕ ИЗ АРШИНА
    def _get_dicts_from_fif(self):

        # ЕСЛИ ID РЕЕСТРА НЕ НАЙДЕН, ИЩЕМ ПО НОМЕРУ В РЕЕСТРЕ
        if not self.mit_search:
            reestr = self.final_dict[self.cur_row_number]['reestr']
            if reestr:
                self.search_thread.url = f"{URL_START}/mit?start=0&rows=100&search={reestr}"
                self.search_thread.run()

        # ЕСЛИ ID РЕЕСТРА НАЙДЕН, ИЩЕМ ИНФОРМАЦИЮ О РЕЕСТРЕ ПО ID
        elif not self.mit and self.mit_search:
            if 'result' in self.mit_search and 'count' in self.mit_search['result']:
                count = self.mit_search['result']['count']

                # ЕСЛИ РЕЗУЛЬТАТОВ ПОИСКА = 1, БЕРЕМ ЕГО ID И ИЩЕМ РЕЕСТР
                if count == 1 and 'items' in self.mit_search['result'] \
                        and 'mit_id' in self.mit_search['result']['items'][0]:
                    mit_id = self.mit_search['result']['items'][0]['mit_id']
                    if mit_id:
                        self.search_thread.url = f"{URL_START}/mit/{mit_id}"
                        self.search_thread.run()

                # ЕСЛИ РЕЗУЛЬТАТОВ ПОИСКА БОЛЬШЕ ОДНОГО И В ТАБЛИЦЕ ЕСТЬ НАИМЕНОВАНИЕ ИЛИ ИЗГОТОВИТЕЛЬ,
                # ИЩЕМ НАИБОЛЬШУЮ ОБЩУЮ ПОДСТРОКУ ПО НАИМЕНОВАНИЮ ИЛИ ЗАВОДУ
                elif count > 1 and 'items' in self.mit_search['result']:
                    items = self.mit_search['result']['items']
                    cur_title = self.final_dict[self.cur_row_number]['title']
                    cur_manufacturer = self.final_dict[self.cur_row_number]['manufacturer']
                    mit_id = ""

                    title_list = []
                    manuf_list = []

                    if cur_title:
                        for item in items:
                            if 'title' in item:
                                title = item['title']
                                match_length = func.get_max_substring(title, cur_title)[0]
                                title_list.append(match_length)
                    if cur_manufacturer:
                        for item in items:
                            if 'manufactorer' in item:
                                manufactorer = item['manufactorer']
                                match_length = func.get_max_substring(manufactorer, cur_manufacturer)[0]
                                manuf_list.append(match_length)
                    if title_list:
                        max_title = max(title_list)
                        max_manuf = 0
                        for i in range(len(title_list)):
                            if title_list[i] == max_title:
                                if manuf_list:
                                    if manuf_list[i] > max_manuf:
                                        max_manuf = manuf_list[i]
                                        if 'mit_id' in self.mit_search['result']['items'][i]:
                                            mit_id = self.mit_search['result']['items'][i]['mit_id']
                                else:
                                    if 'mit_id' in self.mit_search['result']['items'][i]:
                                        mit_id = self.mit_search['result']['items'][i]['mit_id']
                    elif manuf_list:
                        max_manuf = max(manuf_list)
                        for i in range(len(manuf_list)):
                            if manuf_list[i] == max_manuf:
                                if 'mit_id' in self.mit_search['result']['items'][i]:
                                    mit_id = self.mit_search['result']['items'][i]['mit_id']

                    else:
                        if 'mit_id' in self.mit_search['result']['items'][0]:
                            mit_id = self.mit_search['result']['items'][0]['mit_id']

                    # print(mit_id)
                    # print(title_list)
                    # print(manuf_list)

                    if mit_id:
                        self.search_thread.url = f"{URL_START}/mit/{mit_id}"
                        self.search_thread.run()

                    return

    def _on_getting_resp(self, resp):
        if not resp or resp.startswith("Error") or resp.startswith("<!DOCTYPE html>"):
            QMessageBox.critical(self, "Ошибка", f"Возникла ошибка получения сведений из ФГИС \"АРШИН\".\n{resp}")
            return
        else:
            try:
                self.resp_json = json.loads(resp)
            except JSONDecodeError as err:
                QMessageBox.critical(self, "Ошибка", f"Невозможно распознать ответ ФГИС \"АРШИН\".\n{resp}")
                return
            if self.resp_json:
                if "mit?" in self.search_thread.url:
                    self.mit_search = self.resp_json
                    print(self.mit_search)
                    self._get_dicts_from_fif()
                elif "mit/" in self.search_thread.url:
                    self.mit = self.resp_json
                    self._get_dicts_from_fif()

    def _get_mit(self, reestr):
        self.search_thread.url = f"{URL_START}/mit?search={reestr}"
        self.search_thread.run()

    def _verification_start(self, row_number):
        # self._get_dicts_from_fif()

        # import_reg_card_number = self.final_dict[row_number]['reg_card_number']
        # import_reestr = self.final_dict[row_number]['reestr']
        # import_title = self.final_dict[row_number]['title']
        # import_number = self.final_dict[row_number]['number']
        # import_certNum_1 = self.final_dict[row_number]['certNum_1']
        # import_mieta_number_1 = self.final_dict[row_number]['mieta_number_1']

        # # ЕСЛИ ОТСУТСТВУЕТ И ТИП И МОДИФИКАЦИЯ, ИМПОРТ СТРОКИ НЕ ПРОИСХОДИТ
        # if not import_type and not import_modification:

        # for mi_id in self.mi_dict:
        # db_reg_card_number = self.mi_dict[mi_id]['reg_card_number']
        # db_reestr = self.mi_dict[mi_id]['reestr']
        # db_title = self.mi_dict[mi_id]['title']
        # db_modification = self.mi_dict[mi_id]['modification']
        # db_number = self.mi_dict[mi_id]['number']

        # # ЕСЛИ НОМЕР РЕГИСТРАЦИОННОЙ КАРТОЧКИ УЖЕ СУЩЕСТВУЕТ, ПРЕДЛАГАЕМ ПРОПУСТИТЬ ИЛИ СОХРАНИТЬ ПРИБОР
        # # ПОД ДРУГИМ НОМЕРОМ, СОСТОЯЩИМ ИЗ НОМЕРА И ТЕКУЩЕЙ ДАТЫ/ВРЕМЕНИ
        # if db_reg_card_number == import_reg_card_number:
        #     new_reg_card_number = f"{db_reg_card_number} " \
        #                           f"{QDateTime().currentDateTime().toString('dd_MM_yy HH:mm:ss:zzz')}"
        #     dialog = QMessageBox(self)
        #     dialog.setWindowTitle("Дублирование номера регистрационной карточки")
        #     dialog.setText(f"Внимание!\n"
        #                    f"Регистрационная карточка № {import_reg_card_number} уже сохранена в базе.\n"
        #                    f"Она соответствует прибору:\n{db_title} {db_modification} зав. № {db_number}\n"
        #                    f"В импортируемой таблице эта карточка соответствует прибору:\n{import_title} "
        #                    f"{import_modification} зав. № {import_number}\n"
        #                    f"Номер строки в таблице: {row_number}\n\n"
        #                    f"Пропустить импортирование этого прибора или сохранить его под номером "
        #                    f"\'{new_reg_card_number}\' с дальнейшим его изменением пользователем?")
        #     dialog.setIcon(QMessageBox.Question)
        #     btn_save = QPushButton("&Сохранить")
        #     dialog.addButton(btn_save, QMessageBox.AcceptRole)
        #     btn_skip = QPushButton("&Пропустить")
        #     dialog.addButton(btn_skip, QMessageBox.RejectRole)
        #     dialog.setDefaultButton(btn_save)
        #     dialog.setEscapeButton(btn_save)
        #     result = dialog.exec()
        #     if result == 0:
        #         self.final_dict[row_number]['reg_card_number'] = new_reg_card_number
        #         # print(self.mi_dict[mi_id]['reg_card_number'])
        #     else:
        #         del self.final_dict[row_number]
        #         return
        #

        # # ЕСЛИ ПОВТОРЯЕТСЯ НОМЕР РЕЕСТРА И ЗАВОДСКОЙ НОМЕР, ВЫВОДИТСЯ ПРЕДУПРЕЖДЕНИЕ
        # if db_reestr == import_reestr and db_number == import_number:
        #     QMessageBox.information(self, "Внимание!",
        #                             f"Предупреждение по импортируемой строке {row_number}!\n"
        #                             f"Прибор с регистрационным номером типа СИ \'{import_reestr}\' "
        #                             f"и заводским номером \'{import_number}\' уже содержится в базе.\n"
        #                             f"После окончания импорта проверьте форму оборудования и "
        #                             f"удалите дублирующиеся записи, либо запустите автоматическую проверку.")

        cert_numbers = self.final_dict[row_number]['cert_numbers']
        mieta_numbers = self.final_dict[row_number]['mieta_numbers']

        cert_number = False
        mieta_number = False
        for i in range(10):
            if cert_numbers[i]:
                cert_number = True
            if mieta_numbers[i]:
                mieta_number = True
        reestr = self.final_dict[row_number]['reestr']
        title = self.final_dict[row_number]['title']
        number = self.final_dict[row_number]['number']
        inv_number = self.final_dict[row_number]['inv_number']

        if not cert_number and not mieta_number:
            if reestr or title:
                if not number and not inv_number:
                    self.error_row_numbers.append(str(row_number))
                    print("Пропуск - нет номеров")
                    return

            else:
                self.error_row_numbers.append(str(row_number))
                print("Пропуск - все пустое")
                return

        responsiblePerson = func.get_worker_id_from_fio(self.final_dict[row_number]['responsiblePerson'],
                                                        self.worker_dict)
        room = func.get_room_id_from_number(self.final_dict[row_number]['room'], self.room_dict)

        has_manual = str(self.final_dict[row_number]['has_manual']).lower()
        has_pasport = str(self.final_dict[row_number]['has_pasport']).lower()
        has_verif_method = str(self.final_dict[row_number]['has_verif_method']).lower()
        if has_manual == "да" or has_manual == "1" or has_manual == "истина" or has_manual == "true":
            has_manual = 2
        else:
            has_manual = 0
        if has_pasport == "да" or has_pasport == "1" or has_pasport == "истина" or has_pasport == "true":
            has_pasport = 2
        else:
            has_pasport = 0
        if has_verif_method == "да" or has_verif_method == "1" or has_verif_method == "истина" or has_verif_method == "true":
            has_verif_method = 2
        else:
            has_verif_method = 0
        if not self.final_dict[row_number]['reg_card_number']:
            self.final_dict[row_number][
                'reg_card_number'] = f"{QDateTime().currentDateTime().toString('dd_MM_yy HH:mm:ss:zzz')}"

        if self.final_dict[row_number]['reg_card_number'] in self.set_of_card_numbers:
            self.final_dict[row_number][
                'reg_card_number'] = f"{QDateTime().currentDateTime().toString('dd_MM_yy HH:mm:ss:zzz')}"

        self.set_of_card_numbers.add(self.final_dict[row_number]['reg_card_number'])

        sql_insert_mis = f"INSERT INTO mis VALUES (" \
                         f"NULL, " \
                         f"'{self.final_dict[row_number]['reg_card_number']}', " \
                         f"'{self.final_dict[row_number]['measure_code']}', " \
                         f"'СИ', " \
                         f"'{self.final_dict[row_number]['reestr']}', " \
                         f"'{self.final_dict[row_number]['title']}', " \
                         f"'{self.final_dict[row_number]['type']}', " \
                         f"'{self.final_dict[row_number]['modification']}', " \
                         f"'{self.final_dict[row_number]['number']}', " \
                         f"'{self.final_dict[row_number]['inv_number']}', " \
                         f"'{self.final_dict[row_number]['manufacturer']}', " \
                         f"'{self.final_dict[row_number]['manuf_year']}', " \
                         f"'{self.final_dict[row_number]['expl_year']}', " \
                         f"'{self.final_dict[row_number]['diapazon']}', " \
                         f"'{self.final_dict[row_number]['PG']}', " \
                         f"'{self.final_dict[row_number]['KT']}', " \
                         f"'{self.final_dict[row_number]['other_characteristics']}', " \
                         f"'{self.final_dict[row_number]['MPI']}', " \
                         f"'{self.final_dict[row_number]['purpose']}', " \
                         f"{int(responsiblePerson)}, " \
                         f"'{self.final_dict[row_number]['personal']}', " \
                         f"{int(room)}, " \
                         f"'{self.final_dict[row_number]['software_inner']}', " \
                         f"'{self.final_dict[row_number]['software_outer']}', " \
                         f"{int(has_manual)}, " \
                         f"{int(has_pasport)}, " \
                         f"{int(has_verif_method)}, " \
                         f"'{self.final_dict[row_number]['period_TO']}', " \
                         f"'{self.final_dict[row_number]['owner']}', " \
                         f"'{self.final_dict[row_number]['owner_contract']}', " \
                         f"NULL);"
        # print(sql_insert_mis)
        MySQLConnection.verify_connection()
        connection = MySQLConnection.get_connection()
        result = MySQLConnection.execute_query(connection, sql_insert_mis)

        if result[0]:
            # print(row_number)
            mi_id = str(result[1])

        else:
            connection.close()
            return

        if mi_id:
            # ДЕЛИМ ОТДЕЛЫ ЧЕРЕЗ ЗАПЯТУЮ И СОХРАНЯЕМ В БД
            insert_list = self.final_dict[row_number]['department'].split(",")
            insert_set = {func.get_dep_id_from_number(dep_number.strip(), self.dep_dict) for dep_number in insert_list
                          if dep_number}
            insert_set.discard("0")
            if insert_set:
                sql_insert = f"INSERT IGNORE INTO mis_departments VALUES ({int(mi_id)}, " \
                             f"{f'), ({int(mi_id)}, '.join(insert_set)});"
                MySQLConnection.execute_query(connection, sql_insert)

            # СОХРАНЯЕМ ПОВЕРКИ
            vrf_dates = self.final_dict[row_number]['vrf_dates']
            valid_dates = self.final_dict[row_number]['valid_dates']
            results = self.final_dict[row_number]['results']
            organizations = self.final_dict[row_number]['organizations']

            for i in range(10):
                if cert_numbers[i] and mieta_numbers[i]:
                    applicable = results[i].lower()
                    if applicable == "брак" \
                            or applicable == "нет" \
                            or applicable == "0" \
                            or applicable == "ложь" \
                            or applicable == "false":
                        applicable = 0
                    else:
                        applicable = 1
                    for mieta_num in str(mieta_numbers[i]).split(","):
                        mieta_num = mieta_num.strip()
                        sql_insert = f"INSERT INTO mis_vri_info (vri_mi_id, vri_organization, vri_vrfDate, " \
                                     f"vri_validDate, vri_applicable, vri_certNum, vri_mieta_number) VALUES (" \
                                     f"{int(mi_id)}, " \
                                     f"'{organizations[i]}', " \
                                     f"'{vrf_dates[i]}', " \
                                     f"'{valid_dates[i]}', " \
                                     f"{int(applicable)}, " \
                                     f"'{cert_numbers[i]}', " \
                                     f"'{mieta_num}')"

                        MySQLConnection.execute_query(connection, sql_insert)

                elif cert_numbers[i]:
                    applicable = results[i].lower()
                    if applicable == "брак" \
                            or applicable == "нет" \
                            or applicable == "0" \
                            or applicable == "ложь" \
                            or applicable == "false":
                        applicable = 0
                    else:
                        applicable = 1
                    sql_insert = f"INSERT INTO mis_vri_info (vri_mi_id, vri_organization, vri_vrfDate, " \
                                 f"vri_validDate, vri_applicable, vri_certNum) VALUES (" \
                                 f"{int(mi_id)}, " \
                                 f"'{organizations[i]}', " \
                                 f"'{vrf_dates[i]}', " \
                                 f"'{valid_dates[i]}', " \
                                 f"{int(applicable)}, " \
                                 f"'{cert_numbers[i]}')"
                    MySQLConnection.execute_query(connection, sql_insert)

                elif mieta_numbers[i]:
                    for mieta_num in str(mieta_numbers[i]).split(","):
                        mieta_num = mieta_num.strip()
                        sql_insert = f"INSERT INTO mis_vri_info (vri_mi_id, vri_mieta_number) VALUES (" \
                                     f"{int(mi_id)}, " \
                                     f"'{mieta_num}');"
                        MySQLConnection.execute_query(connection, sql_insert)

        connection.close()

    def _format_string(self, string):
        new_str = string
        if "\"" in new_str:
            new_str = new_str.replace("\"", "\\\"")
        if "\'" in new_str:
            new_str = new_str.replace("\'", "\\\'")
        if "\r" in new_str:
            new_str = new_str.replace("\r", "")
        if "_x000D_" in new_str:
            new_str = new_str.replace("_x000D_", " ")
        return new_str


if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    window = EquipmentImportFileWidget()

    window.show()
    sys.exit(app.exec_())
