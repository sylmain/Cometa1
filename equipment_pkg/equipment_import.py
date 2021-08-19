import json
from json.decoder import JSONDecodeError

from PyQt5.QtCore import Qt, QSettings, QUrl, QThread, pyqtSignal
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import QWidget, QFileDialog, QMessageBox
from openpyxl import load_workbook

from equipment_pkg.ui_equipment_import import Ui_Form
from functions_pkg.send_get_request import GetRequest
from global_vars import Globals

SETTINGS = QSettings(Globals.settings_path_string, QSettings.IniFormat)
SETTINGS.setIniCodec("UTF-8")
URL_START = "https://fgis.gost.ru/fundmetrology/eapi"


class SearchThread(QThread):
    msg_signal = pyqtSignal(str)

    def __init__(self, parent=None):
        QThread.__init__(self, parent)
        self.url = ""
        self.is_running = True

    def run(self):
        if self.is_running:
            self.sleep(1)
            print("thread running")
            print(f" {self.url}")
            resp = GetRequest.getRequest(self.url)
            print(f"  {resp}")
            print("    thread stopped")
            self.msg_signal.emit(resp)
        else:
            self.msg_signal.emit("stop")


class EquipmentImportWidget(QWidget):

    def __init__(self, parent=None):
        super(EquipmentImportWidget, self).__init__(parent)
        # self.setWindowModality(Qt.WindowModal)
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.list_of_numbers = list()
        self.numbers_model = QStandardItemModel(0, 5, parent=self)

        self.ui.tableView_numbers.setModel(self.numbers_model)

        self.search_thread = SearchThread()

        self.number_vri_dict = dict()

        self.numbers_model.setHorizontalHeaderLabels(
            ["Номер документа", "Номер реестра", "Наименование", "Тип", "Заводской номер"])
        self.ui.tableView_numbers.setColumnWidth(0, 120)
        self.ui.tableView_numbers.setColumnWidth(0, 100)
        self.ui.tableView_numbers.setColumnWidth(1, 200)
        self.ui.tableView_numbers.setColumnWidth(2, 100)
        self.ui.tableView_numbers.setColumnWidth(3, 100)

        self._create_connects()

    def _create_connects(self):
        self.ui.pushButton_file_select.clicked.connect(self._file_select)
        self.search_thread.msg_signal.connect(self._on_getting_resp, Qt.QueuedConnection)

    def _file_select(self):
        dir_name = QUrl(str(SETTINGS.value("paths/cometa_path")))
        file = QFileDialog.getOpenFileUrl(caption="Выберите файл с номерами свидетельств", filter="Excel (*.xlsx)")
        file_path = file[0].path()[1:]
        wb = load_workbook(filename=file_path)
        sheet = wb['Лист1']
        count = sheet.max_row
        for i in range(1, count + 1):
            number = str(sheet[f"A{i}"].value)
            self.list_of_numbers.append(number)
        self._send_get()

    def _send_get(self):
        if self.list_of_numbers:
            self.number = self.list_of_numbers.pop(0)
            self.search_thread.url = f"{URL_START}/vri/1-{self.number}"
            self.search_thread.run()
        else:
            QMessageBox.information(self, "Завершено", "Поиск завершен")
            for vri in self.number_vri_dict:
                print(f"{vri} : {self.number_vri_dict[vri]}")

    def _on_getting_resp(self, resp):
        if not resp or resp.startswith("Error") or resp.startswith("<!DOCTYPE html>"):
            QMessageBox.critical(self, "Ошибка", f"Возникла ошибка получения сведений из ФГИС \"АРШИН\".\n{resp}")
            return
        else:
            try:
                self.resp_json = json.loads(resp)
            except JSONDecodeError as err:
                QMessageBox.critical(self, "Ошибка", f"Невозможно распознать ответ ФГИС \"АРШИН\".\n{resp}")
                return
            if self.resp_json:
                self.number_vri_dict[self.number] = self.resp_json
                self._fill_table()
                self._send_get()

    def _fill_table(self):
        row = list()
        if 'vriInfo' in self.resp_json['result']:
            vriInfo = self.resp_json['result']['vriInfo']
            if 'applicable' in vriInfo and 'certNum' in vriInfo['applicable']:
                cert_number = vriInfo['applicable']['certNum']
            elif 'inapplicable' in vriInfo and 'noticeNum' in vriInfo['applicable']:
                cert_number = vriInfo['applicable']['noticeNum']
            row.append(QStandardItem(cert_number))
        if 'result' in self.resp_json and 'miInfo' in self.resp_json['result']:
            if 'etaMI' in self.resp_json['result']['miInfo']:
                miInfo = self.resp_json['result']['miInfo']['etaMI']
            elif 'singleMI' in self.resp_json['result']['miInfo']:
                miInfo = self.resp_json['result']['miInfo']['singleMI']
            elif 'partyMI' in self.resp_json['result']['miInfo']:
                miInfo = self.resp_json['result']['miInfo']['partyMI']

            if 'mitypeNumber' in miInfo:
                row.append(QStandardItem(miInfo['mitypeNumber']))
            if 'mitypeTitle' in miInfo:
                row.append(QStandardItem(miInfo['mitypeTitle']))
            if 'modification' in miInfo:
                row.append(QStandardItem(miInfo['modification']))
            if 'manufactureNum' in miInfo:
                row.append(QStandardItem(miInfo['manufactureNum']))

        if row:
            self.numbers_model.appendRow(row)
        self.ui.tableView_numbers.resizeRowsToContents()

