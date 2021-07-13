import mysql.connector
import requests
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtWidgets import QMainWindow, QApplication
from mysql.connector import Error
from openpyxl import load_workbook

from update_pkg.ui_update import Ui_MainWindow


class UpdateThread(QThread):
    msg_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)

    def __init__(self, parent=None):
        QThread.__init__(self, parent)
        self.running = False
        wb = load_workbook(filename='./fif_results.xlsx')
        self.sheet = wb['Sheet1']
        self.count = self.sheet.max_row - 3

    def run(self):
        self.running = True
        numbers = set()
        with open("./certNumbers.txt", "r", encoding="utf-8") as f:
            for line in f:
                numbers.add(line.strip()[-8:])
            f.close()

        x = 4

        try:
            connection = mysql.connector.connect(
                host="csmsev.ru",
                port="3306",
                user="sylmain",
                passwd="muahiwwana",
                database="sylmain_job"
            )
        except Error as e:
            self.error_signal.emit(f"error: {e}")
            return
        while self.running and x < self.sheet.max_row + 1:
            cell = 'K' + str(x)
            vri_id = self.sheet[cell].value[-8:]
            if vri_id not in numbers:
                try:
                    self.sleep(1)
                    resp = requests.get(f"https://fgis.gost.ru/fundmetrology/eapi/vri/1-{vri_id}")
                    json_resp = resp.json()
                    if "result" in json_resp:
                        if "etaMI" in json_resp['result']['miInfo']:
                            reestr = json_resp['result']['miInfo']['etaMI']['mitypeNumber']
                            title = json_resp['result']['miInfo']['etaMI']['mitypeTitle']
                            modification = json_resp['result']['miInfo']['etaMI']['modification']
                            manufactureNum = json_resp['result']['miInfo']['etaMI']['manufactureNum']

                        elif "singleMI" in json_resp['result']['miInfo']:
                            if "mitypeNumber" in json_resp['result']['miInfo']['singleMI']:
                                reestr = json_resp['result']['miInfo']['singleMI']['mitypeNumber']
                                title = json_resp['result']['miInfo']['singleMI']['mitypeTitle']
                            else:
                                reestr = "СГ101-32"
                                title = json_resp['result']['miInfo']['singleMI']['mitypeTitle'].replace("СГ101-32 - ",
                                                                                                         "")
                                # print(title)
                            modification = json_resp['result']['miInfo']['singleMI']['modification']
                            manufactureNum = json_resp['result']['miInfo']['singleMI']['manufactureNum']

                        elif "partyMI" in json_resp['result']['miInfo']:
                            reestr = json_resp['result']['miInfo']['partyMI']['mitypeNumber']
                            title = json_resp['result']['miInfo']['partyMI']['mitypeTitle']
                            modification = json_resp['result']['miInfo']['partyMI']['modification']
                            manufactureNum = json_resp['result']['miInfo']['partyMI']['quantity']

                        if "applicable" in json_resp['result']['vriInfo']:
                            certNum = json_resp['result']['vriInfo']['applicable']['certNum']
                            if "stickerNum" in json_resp['result']['vriInfo']['applicable']:
                                sticker = json_resp['result']['vriInfo']['applicable']['stickerNum']
                                if sticker == "Нет данных":
                                    sticker = "-"
                            else:
                                sticker = "-"
                            valid = json_resp['result']['vriInfo']['validDate']
                            validDate = f"{valid.split('.')[2]}-{valid.split('.')[1]}-{valid.split('.')[0]}"
                            result = "ГОДЕН"
                        elif "inapplicable" in json_resp['result']['vriInfo']:
                            certNum = json_resp['result']['vriInfo']['inapplicable']['noticeNum']
                            sticker = "-"
                            validDate = "0000-00-00"
                            result = "БРАК"
                        vrf = json_resp['result']['vriInfo']['vrfDate']
                        vrfDate = f"{vrf.split('.')[2]}-{vrf.split('.')[1]}-{vrf.split('.')[0]}"
                        miOwner = json_resp['result']['vriInfo']['miOwner']
                        miOwner = miOwner.replace("«", "\"")
                        miOwner = miOwner.replace("»", "\"")
                    else:
                        print("некорректный JSON")
                        self.error_signal.emit(f"bad JSON")
                        return
                    query = f"INSERT IGNORE INTO complete_jobs2021_2 VALUES " \
                            f"('{title}', " \
                            f"'{modification}', " \
                            f"'{miOwner}', " \
                            f"'{manufactureNum}', " \
                            f"'{reestr}', " \
                            f"'{certNum}', " \
                            f"'{vrfDate}', " \
                            f"'{sticker}', " \
                            f"'{result}', " \
                            f"'{validDate}', " \
                            f"'')"
                    cursor = connection.cursor()
                    cursor.execute(query)
                    connection.commit()
                    with open("./certNumbers.txt", "a", encoding="utf-8") as f:
                        f.write(f"{certNum}\n")
                except Error as e:
                    self.error_signal.emit(f"error: {e}")
                    return
            self.msg_signal.emit(f"{x - 3}/{self.count}")
            x += 1
            self.progress_signal.emit(x)
        self.msg_signal.emit(f"Выполняю обновление кода заказчика")
        query = f"UPDATE " \
                f"  complete_jobs2021_2, " \
                f"  customers " \
                f"SET " \
                f"  complete_jobs2021_2.code = customers.code " \
                f"WHERE " \
                f"  complete_jobs2021_2.code = '' AND complete_jobs2021_2.customer = customers.name"
        cursor = connection.cursor()
        cursor.execute(query)
        connection.commit()
        connection.close()
        self.msg_signal.emit(f"Все операции завершены")


class Update(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(Update, self).__init__()

        print(ord("»"))
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.update_thread = UpdateThread()
        self.ui.progressBar.setRange(0, self.update_thread.count)

        self.update_thread.msg_signal.connect(self.on_change, Qt.QueuedConnection)
        self.update_thread.error_signal.connect(self.on_error, Qt.QueuedConnection)
        self.update_thread.progress_signal.connect(self.on_progress, Qt.QueuedConnection)

        self.update_thread.started.connect(lambda: print("Поток запущен"))
        self.update_thread.finished.connect(lambda: print("Поток остановлен"))

        self.ui.pushButton_start.clicked.connect(self.on_start)
        self.ui.pushButton_stop.clicked.connect(self.on_stop)

    def on_start(self):
        if not self.update_thread.isRunning():
            self.ui.label_counter.setText("Стартуем...")
            self.ui.label_error.setText("")
            self.ui.progressBar.setValue(0)

            self.ui.pushButton_start.setEnabled(False)
            self.ui.pushButton_stop.setEnabled(True)
            self.update_thread.start()

    def on_stop(self):
        self.update_thread.running = False
        self.ui.pushButton_start.setEnabled(True)
        self.ui.pushButton_stop.setEnabled(False)

    def on_change(self, s):
        self.ui.label_counter.setText(s)

    def on_error(self, s):
        self.ui.pushButton_start.setEnabled(True)
        self.ui.pushButton_stop.setEnabled(False)
        self.ui.label_error.setText(s)

    def on_progress(self, i):
        self.ui.progressBar.setValue(i)


if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    mainWindow = Update()
    mainWindow.show()
    sys.exit(app.exec_())
